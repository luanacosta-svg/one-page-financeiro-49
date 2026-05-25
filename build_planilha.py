"""
Gera Planilha_Resultados_2026.xlsx aplicando o 49Pay Design System.

Paleta:
- slate-800 #1E293B  (cabeçalhos de aba)
- slate-700 #334155  (texto)
- slate-200 #E2E8F0  (bordas)
- slate-100 #F1F5F9  (totais)
- slate-50  #F8FAFC  (fundo)
- orange-500 #F97316 (destaque)
- orange-600 #EA580C (texto auto)
- orange-100 #FFEDD5 (sub-headers)
- orange-50  #FFF7ED (células auto)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============ TOKENS ============
NAVY    = '1E293B'  # slate-800
TXT     = '334155'  # slate-700
BORDER  = 'E2E8F0'  # slate-200
SUBTLE  = 'F1F5F9'  # slate-100
BG      = 'F8FAFC'  # slate-50
ORANGE  = 'F97316'  # orange-500
ORG_DK  = 'EA580C'  # orange-600
ORG_LT  = 'FFEDD5'  # orange-100
ORG_BG  = 'FFF7ED'  # orange-50
AMBER_BG = 'FEF3C7' # amber-100
AMBER_TXT = 'B45309' # amber-700

FONT_NAME = 'Inter'  # graceful-degrades pra system em Sheets

# Borders
thin_side = Side(style='thin', color=BORDER)
box_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Styles helpers
def navy_header(text_size=11):
    return {
        'font': Font(name=FONT_NAME, size=text_size, bold=True, color='FFFFFF'),
        'fill': PatternFill('solid', fgColor=NAVY),
        'alignment': Alignment(horizontal='left', vertical='center', indent=1),
        'border': box_border,
    }

def orange_subheader():
    return {
        'font': Font(name=FONT_NAME, size=10, bold=True, color=ORG_DK),
        'fill': PatternFill('solid', fgColor=ORG_LT),
        'alignment': Alignment(horizontal='left', vertical='center', indent=1),
        'border': box_border,
    }

def label_cell():
    return {
        'font': Font(name=FONT_NAME, size=10, color=TXT),
        'alignment': Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True),
        'border': box_border,
    }

def input_cell(num_format='#,##0;(#,##0);"-"'):
    return {
        'font': Font(name=FONT_NAME, size=10, color=TXT),
        'fill': PatternFill('solid', fgColor='FFFFFF'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': box_border,
        'number_format': num_format,
    }

def auto_cell(num_format='#,##0;(#,##0);"-"'):
    return {
        'font': Font(name=FONT_NAME, size=10, italic=True, color=ORG_DK),
        'fill': PatternFill('solid', fgColor=ORG_BG),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': box_border,
        'number_format': num_format,
    }

def total_cell(num_format='#,##0;(#,##0);"-"'):
    return {
        'font': Font(name=FONT_NAME, size=10, bold=True, color=NAVY),
        'fill': PatternFill('solid', fgColor=SUBTLE),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': box_border,
        'number_format': num_format,
    }

def column_header():
    return {
        'font': Font(name=FONT_NAME, size=9, bold=True, color=TXT),
        'fill': PatternFill('solid', fgColor=SUBTLE),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': box_border,
    }

CURRENCY = '"R$" #,##0;("R$" #,##0);"-"'
INT      = '#,##0;(#,##0);"-"'
PCT      = '0.0%;(0.0%);"-"'
DEC      = '0.0;(0.0);"-"'

# ============ HELPERS ============
def apply_style(cell, style):
    for k, v in style.items():
        setattr(cell, k, v)

def write_cell(ws, row, col, value, style=None):
    c = ws.cell(row=row, column=col, value=value)
    if style:
        apply_style(c, style)
    return c

def set_widths(ws, widths):
    """widths: dict {col_letter: width}"""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ============ BUILD ============
wb = Workbook()
wb.remove(wb.active)

MONTHS = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

# =============================================================
# ABA 1 — INSTRUÇÕES
# =============================================================
ws = wb.create_sheet('📋 Instruções')
ws.sheet_view.showGridLines = False

# Hero
ws.merge_cells('B2:G2')
ws['B2'] = 'PLANILHA DE RESULTADOS · 49 Educação'
apply_style(ws['B2'], {
    'font': Font(name=FONT_NAME, size=22, bold=True, color=NAVY),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
set_row_height(ws, 2, 40)

ws.merge_cells('B3:G3')
ws['B3'] = 'Resultado mensal consolidado · 3 áreas de receita + Financeiro'
apply_style(ws['B3'], {
    'font': Font(name=FONT_NAME, size=11, color=TXT),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})

# Como funciona
ws.merge_cells('B5:G5')
ws['B5'] = '  COMO FUNCIONA'
apply_style(ws['B5'], navy_header(11))
set_row_height(ws, 5, 28)

ws.merge_cells('B6:G6')
ws['B6'] = ('Cada líder preenche apenas a sua aba até o dia 15 do mês seguinte. '
            'A Luana consolida o Financeiro e o dashboard HTML lê a aba "📤 Export" automaticamente.')
apply_style(ws['B6'], {
    'font': Font(name=FONT_NAME, size=10, color=TXT),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True),
})
set_row_height(ws, 6, 36)

# Responsáveis
ws.merge_cells('B8:G8')
ws['B8'] = '  ÁREAS E RESPONSÁVEIS'
apply_style(ws['B8'], navy_header(11))
set_row_height(ws, 8, 28)

resp_headers = ['Aba', 'Responsável', 'Prazo de Envio']
for i, h in enumerate(resp_headers):
    write_cell(ws, 9, 2 + i*2, h, column_header())
    ws.merge_cells(start_row=9, start_column=2+i*2, end_row=9, end_column=3+i*2)

resp_rows = [
    ('🛠 Produto', 'Tiago Toigo', 'Até dia 15 do mês seguinte'),
    ('🤝 Comercial B2B', 'Marcelo Câmara', 'Até dia 15 do mês seguinte'),
    ('📋 Licitações', 'Rafael Peck', 'Até dia 15 do mês seguinte'),
    ('💰 Financeiro', 'Luana / Financeiro', 'Após todas as áreas preencherem'),
    ('📤 Export', 'Automático (consolidação)', 'Lido pelo One Page HTML'),
]
for i, (a, b, c) in enumerate(resp_rows):
    r = 10 + i
    write_cell(ws, r, 2, a, label_cell())
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    write_cell(ws, r, 4, b, label_cell())
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    write_cell(ws, r, 6, c, label_cell())
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    set_row_height(ws, r, 22)

# Legenda
ws.merge_cells('B17:G17')
ws['B17'] = '  LEGENDA DE CORES'
apply_style(ws['B17'], navy_header(11))
set_row_height(ws, 17, 28)

legenda = [
    ('Campo de preenchimento', 'Preencha o número conforme rótulo do indicador', 'input'),
    ('Calculado automaticamente', 'Não edite — fórmula puxa de outras células', 'auto'),
    ('Total / Cabeçalho de seção', 'Linhas de subtotal e categorias', 'total'),
]
for i, (titulo, desc, tipo) in enumerate(legenda):
    r = 18 + i
    # swatch
    cell = ws.cell(row=r, column=2, value='')
    if tipo == 'input':
        apply_style(cell, {'fill': PatternFill('solid', fgColor='FFFFFF'), 'border': box_border})
    elif tipo == 'auto':
        apply_style(cell, auto_cell())
        cell.value = '1.000'
    else:
        apply_style(cell, total_cell())
        cell.value = '1.000'
    write_cell(ws, r, 3, titulo, {**label_cell(), 'font': Font(name=FONT_NAME, size=10, bold=True, color=NAVY)})
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    write_cell(ws, r, 5, desc, label_cell())
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    set_row_height(ws, r, 22)

# Importante
ws.merge_cells('B23:G23')
ws['B23'] = '  ⚠ IMPORTANTE — VENDIDO vs RECEBIDO'
apply_style(ws['B23'], {
    'font': Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF'),
    'fill': PatternFill('solid', fgColor=ORANGE),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
    'border': box_border,
})
set_row_height(ws, 23, 28)

ws.merge_cells('B24:G25')
ws['B24'] = ('Os líderes preenchem o que foi VENDIDO no mês (contratos fechados). '
             'A Luana preenche o que foi RECEBIDO (entrada em caixa). '
             'O dashboard mostra o gap entre vendido e recebido por área.')
apply_style(ws['B24'], {
    'font': Font(name=FONT_NAME, size=10, color=TXT),
    'fill': PatternFill('solid', fgColor=ORG_BG),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True),
    'border': box_border,
})
set_row_height(ws, 24, 24)
set_row_height(ws, 25, 24)

set_widths(ws, {'A': 2, 'B': 22, 'C': 18, 'D': 22, 'E': 14, 'F': 16, 'G': 20, 'H': 2})

# =============================================================
# HELPER pra abas operacionais
# =============================================================
def build_area_sheet(ws, title, subtitle, lead, sections):
    """sections: list of (group_name, [(label, formula_or_None, num_format)])"""
    ws.sheet_view.showGridLines = False

    # Hero
    ws.merge_cells('B2:O2')
    ws['B2'] = f'{title} — {lead}'
    apply_style(ws['B2'], {
        'font': Font(name=FONT_NAME, size=18, bold=True, color=NAVY),
        'alignment': Alignment(horizontal='left', vertical='center', indent=1),
    })
    set_row_height(ws, 2, 36)

    ws.merge_cells('B3:O3')
    ws['B3'] = subtitle
    apply_style(ws['B3'], {
        'font': Font(name=FONT_NAME, size=10, color=TXT),
        'alignment': Alignment(horizontal='left', vertical='center', indent=1),
    })

    # Header row (R5)
    headers = ['Indicador'] + MONTHS + ['Total/Média']
    for i, h in enumerate(headers):
        c = write_cell(ws, 5, 2 + i, h, column_header())
        ws.column_dimensions[get_column_letter(2 + i)].width = 32 if i == 0 else 11
    set_row_height(ws, 5, 28)

    current_row = 6
    metric_rows = []  # collect for later if needed

    for group_name, metrics in sections:
        # Section sub-header
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=15)
        write_cell(ws, current_row, 2, f'  {group_name}', orange_subheader())
        set_row_height(ws, current_row, 24)
        current_row += 1

        for metric in metrics:
            label, formula_dict, fmt = metric
            # label
            write_cell(ws, current_row, 2, label, label_cell())
            # 12 months
            for m in range(12):
                col = 3 + m
                col_letter = get_column_letter(col)
                if formula_dict is not None:
                    # formula per column — formula_dict can be a callable or template
                    if callable(formula_dict):
                        formula = formula_dict(col_letter, current_row)
                    else:
                        formula = formula_dict.format(col=col_letter, row=current_row)
                    write_cell(ws, current_row, col, formula, auto_cell(fmt))
                else:
                    write_cell(ws, current_row, col, None, input_cell(fmt))
            # Total/Average (col 15 = O)
            row_letter_start = 'C'
            row_letter_end = 'N'
            if fmt == PCT or fmt == DEC:
                total_formula = f'=IFERROR(AVERAGE({row_letter_start}{current_row}:{row_letter_end}{current_row}),0)'
            else:
                total_formula = f'=SUM({row_letter_start}{current_row}:{row_letter_end}{current_row})'
            write_cell(ws, current_row, 15, total_formula, total_cell(fmt))
            metric_rows.append((label, current_row))
            set_row_height(ws, current_row, 22)
            current_row += 1

        # Spacer
        current_row += 1

    # Freeze C6 (header row + indicator column)
    ws.freeze_panes = 'C6'

    return metric_rows

# =============================================================
# ABA 2 — PRODUTO
# =============================================================
ws = wb.create_sheet('🛠 Produto')
build_area_sheet(ws, '🛠 PRODUTO', 'Entrega das aulas/hot seats, saúde da base contratada, vendas de 49 Code e tech', 'Tiago Toigo', [
    ('ENTREGA', [
        ('Aulas / Hot Seats Realizadas', None, INT),
        ('Marcos Entregues no Mês', None, INT),
        ('SLA de Entrega (%)', None, PCT),
    ]),
    ('SAÚDE DA BASE', [
        ('Contratos Ativos', None, INT),
        ('NPS', None, INT),
    ]),
    ('RECEITA (VENDIDA NO MÊS)', [
        ('Vendas de 49 Code (qtd)', None, INT),
        ('Receita de 49 Code (R$)', None, CURRENCY),
        ('Receita Total da Área (R$)', None, CURRENCY),
    ]),
    ('TECH', [
        ('Uptime das Plataformas (%)', None, PCT),
        ('Bugs Críticos Reportados', None, INT),
        ('Bugs Críticos Resolvidos', None, INT),
        ('Features Lançadas no Mês', None, INT),
    ]),
])

# =============================================================
# ABA 3 — B2B
# =============================================================
ws = wb.create_sheet('🤝 B2B')

# Para o B2B precisamos das fórmulas auto referenciando linhas específicas.
# Vou construir manualmente já que tem fórmulas que referenciam outras linhas.
ws.sheet_view.showGridLines = False
ws.merge_cells('B2:O2')
ws['B2'] = '🤝 COMERCIAL B2B — Marcelo Câmara'
apply_style(ws['B2'], {
    'font': Font(name=FONT_NAME, size=18, bold=True, color=NAVY),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
set_row_height(ws, 2, 36)
ws.merge_cells('B3:O3')
ws['B3'] = 'Captação de novos clientes B2B (Sebrae, SATC, PUC, etc), funil e fechamento'
apply_style(ws['B3'], {
    'font': Font(name=FONT_NAME, size=10, color=TXT),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
headers = ['Indicador'] + MONTHS + ['Total/Média']
for i, h in enumerate(headers):
    write_cell(ws, 5, 2 + i, h, column_header())
    ws.column_dimensions[get_column_letter(2 + i)].width = 32 if i == 0 else 11
set_row_height(ws, 5, 28)

# Sub-header MARKETING
ws.merge_cells('B6:O6')
write_cell(ws, 6, 2, '  MARKETING / FUNIL', orange_subheader())
set_row_height(ws, 6, 24)
# 7: ADS (input)
write_cell(ws, 7, 2, 'Investimento ADS (R$)', label_cell())
for m in range(12):
    write_cell(ws, 7, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 7, 15, '=SUM(C7:N7)', total_cell(CURRENCY))
set_row_height(ws, 7, 22)
# 8: Leads
write_cell(ws, 8, 2, 'Leads Gerados', label_cell())
for m in range(12):
    write_cell(ws, 8, 3 + m, None, input_cell(INT))
write_cell(ws, 8, 15, '=SUM(C8:N8)', total_cell(INT))
set_row_height(ws, 8, 22)
# 9: MQLs
write_cell(ws, 9, 2, 'MQLs', label_cell())
for m in range(12):
    write_cell(ws, 9, 3 + m, None, input_cell(INT))
write_cell(ws, 9, 15, '=SUM(C9:N9)', total_cell(INT))
set_row_height(ws, 9, 22)
# 10: SQLs
write_cell(ws, 10, 2, 'SQLs', label_cell())
for m in range(12):
    write_cell(ws, 10, 3 + m, None, input_cell(INT))
write_cell(ws, 10, 15, '=SUM(C10:N10)', total_cell(INT))
set_row_height(ws, 10, 22)
# 11: CAC (auto) = ADS / Contratos Fechados (linha 17)
write_cell(ws, 11, 2, 'CAC (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 11, 3 + m, f'=IFERROR({col}7/{col}17,0)', auto_cell(CURRENCY))
write_cell(ws, 11, 15, '=IFERROR(SUM(C7:N7)/SUM(C17:N17),0)', total_cell(CURRENCY))
set_row_height(ws, 11, 22)

# Sub-header PIPELINE
ws.merge_cells('B13:O13')
write_cell(ws, 13, 2, '  PIPELINE & VENDAS', orange_subheader())
set_row_height(ws, 13, 24)
# 14: Pipeline ativo
write_cell(ws, 14, 2, 'Pipeline Ativo (R$)', label_cell())
for m in range(12):
    write_cell(ws, 14, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 14, 15, '=SUM(C14:N14)', total_cell(CURRENCY))
set_row_height(ws, 14, 22)
# 15: Novos leads B2B no mês
write_cell(ws, 15, 2, 'Novos Leads B2B no Mês', label_cell())
for m in range(12):
    write_cell(ws, 15, 3 + m, None, input_cell(INT))
write_cell(ws, 15, 15, '=SUM(C15:N15)', total_cell(INT))
set_row_height(ws, 15, 22)
# 16: Propostas
write_cell(ws, 16, 2, 'Propostas Enviadas', label_cell())
for m in range(12):
    write_cell(ws, 16, 3 + m, None, input_cell(INT))
write_cell(ws, 16, 15, '=SUM(C16:N16)', total_cell(INT))
set_row_height(ws, 16, 22)
# 17: Contratos
write_cell(ws, 17, 2, 'Contratos Fechados (qtd)', label_cell())
for m in range(12):
    write_cell(ws, 17, 3 + m, None, input_cell(INT))
write_cell(ws, 17, 15, '=SUM(C17:N17)', total_cell(INT))
set_row_height(ws, 17, 22)
# 18: Receita B2B Vendida
write_cell(ws, 18, 2, 'Receita B2B Vendida (R$)', label_cell())
for m in range(12):
    write_cell(ws, 18, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 18, 15, '=SUM(C18:N18)', total_cell(CURRENCY))
set_row_height(ws, 18, 22)
# 19: Ticket médio (auto) = Receita / Contratos
write_cell(ws, 19, 2, 'Ticket Médio B2B (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 19, 3 + m, f'=IFERROR({col}18/{col}17,0)', auto_cell(CURRENCY))
write_cell(ws, 19, 15, '=IFERROR(SUM(C18:N18)/SUM(C17:N17),0)', total_cell(CURRENCY))
set_row_height(ws, 19, 22)
# 20: Taxa de fechamento (auto) = Contratos / SQLs
write_cell(ws, 20, 2, 'Taxa de Fechamento (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 20, 3 + m, f'=IFERROR({col}17/{col}10,0)', auto_cell(PCT))
write_cell(ws, 20, 15, '=IFERROR(AVERAGE(C20:N20),0)', total_cell(PCT))
set_row_height(ws, 20, 22)
# 21: Ciclo médio
write_cell(ws, 21, 2, 'Ciclo Médio de Venda (dias)', label_cell())
for m in range(12):
    write_cell(ws, 21, 3 + m, None, input_cell(INT))
write_cell(ws, 21, 15, '=IFERROR(AVERAGE(C21:N21),0)', total_cell(INT))
set_row_height(ws, 21, 22)

ws.freeze_panes = 'C6'

# =============================================================
# ABA 4 — LICITAÇÕES
# =============================================================
ws = wb.create_sheet('📋 Licitações')
ws.sheet_view.showGridLines = False
ws.merge_cells('B2:O2')
ws['B2'] = '📋 LICITAÇÕES — Rafael Peck'
apply_style(ws['B2'], {
    'font': Font(name=FONT_NAME, size=18, bold=True, color=NAVY),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
set_row_height(ws, 2, 36)
ws.merge_cells('B3:O3')
ws['B3'] = 'Receita via editais públicos · pipeline, prestação de contas e operacional'
apply_style(ws['B3'], {
    'font': Font(name=FONT_NAME, size=10, color=TXT),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
headers = ['Indicador'] + MONTHS + ['Total/Média']
for i, h in enumerate(headers):
    write_cell(ws, 5, 2 + i, h, column_header())
    ws.column_dimensions[get_column_letter(2 + i)].width = 32 if i == 0 else 11
set_row_height(ws, 5, 28)

# PIPELINE
ws.merge_cells('B6:O6')
write_cell(ws, 6, 2, '  PIPELINE DE EDITAIS', orange_subheader())
set_row_height(ws, 6, 24)
write_cell(ws, 7, 2, 'Editais Monitorados', label_cell())
for m in range(12): write_cell(ws, 7, 3 + m, None, input_cell(INT))
write_cell(ws, 7, 15, '=SUM(C7:N7)', total_cell(INT))
set_row_height(ws, 7, 22)
write_cell(ws, 8, 2, 'Editais Submetidos no Mês', label_cell())
for m in range(12): write_cell(ws, 8, 3 + m, None, input_cell(INT))
write_cell(ws, 8, 15, '=SUM(C8:N8)', total_cell(INT))
set_row_height(ws, 8, 22)
write_cell(ws, 9, 2, 'Editais Ganhos no Mês', label_cell())
for m in range(12): write_cell(ws, 9, 3 + m, None, input_cell(INT))
write_cell(ws, 9, 15, '=SUM(C9:N9)', total_cell(INT))
set_row_height(ws, 9, 22)
# 10: Taxa de êxito auto = Ganhos / Submetidos
write_cell(ws, 10, 2, 'Taxa de Êxito (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 10, 3 + m, f'=IFERROR({col}9/{col}8,0)', auto_cell(PCT))
write_cell(ws, 10, 15, '=IFERROR(AVERAGE(C10:N10),0)', total_cell(PCT))
set_row_height(ws, 10, 22)

# RECEITA
ws.merge_cells('B12:O12')
write_cell(ws, 12, 2, '  RECEITA (VENDIDA NO MÊS)', orange_subheader())
set_row_height(ws, 12, 24)
write_cell(ws, 13, 2, 'Receita de Editais Ganhos (R$)', label_cell())
for m in range(12): write_cell(ws, 13, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 13, 15, '=SUM(C13:N13)', total_cell(CURRENCY))
set_row_height(ws, 13, 22)
write_cell(ws, 14, 2, 'Valor Total Pipeline de Editais (R$)', label_cell())
for m in range(12): write_cell(ws, 14, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 14, 15, '=SUM(C14:N14)', total_cell(CURRENCY))
set_row_height(ws, 14, 22)

# PRESTAÇÃO DE CONTAS
ws.merge_cells('B16:O16')
write_cell(ws, 16, 2, '  PRESTAÇÃO DE CONTAS', orange_subheader())
set_row_height(ws, 16, 24)
for i, lbl in enumerate(['Contratos Vigentes em Execução',
                          'Prestações de Contas Entregues',
                          'Prestações Pendentes',
                          'Prestações em Atraso']):
    r = 17 + i
    write_cell(ws, r, 2, lbl, label_cell())
    for m in range(12): write_cell(ws, r, 3 + m, None, input_cell(INT))
    write_cell(ws, r, 15, f'=SUM(C{r}:N{r})', total_cell(INT))
    set_row_height(ws, r, 22)

# OPERACIONAL
ws.merge_cells('B22:O22')
write_cell(ws, 22, 2, '  OPERACIONAL', orange_subheader())
set_row_height(ws, 22, 24)
for i, lbl in enumerate(['TRs em Aprovação', 'Follow-ups Agendados']):
    r = 23 + i
    write_cell(ws, r, 2, lbl, label_cell())
    for m in range(12): write_cell(ws, r, 3 + m, None, input_cell(INT))
    write_cell(ws, r, 15, f'=SUM(C{r}:N{r})', total_cell(INT))
    set_row_height(ws, r, 22)

ws.freeze_panes = 'C6'

# =============================================================
# ABA 5 — FINANCEIRO (Luana)
# =============================================================
ws = wb.create_sheet('💰 Financeiro')
ws.sheet_view.showGridLines = False
ws.merge_cells('B2:O2')
ws['B2'] = '💰 FINANCEIRO — Luana / Financeiro'
apply_style(ws['B2'], {
    'font': Font(name=FONT_NAME, size=18, bold=True, color=NAVY),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
set_row_height(ws, 2, 36)
ws.merge_cells('B3:O3')
ws['B3'] = 'Caixa, recebimentos efetivos, despesas e indicadores · preenchido após áreas'
apply_style(ws['B3'], {
    'font': Font(name=FONT_NAME, size=10, color=TXT),
    'alignment': Alignment(horizontal='left', vertical='center', indent=1),
})
headers = ['Indicador'] + MONTHS + ['Total/Média']
for i, h in enumerate(headers):
    write_cell(ws, 5, 2 + i, h, column_header())
    ws.column_dimensions[get_column_letter(2 + i)].width = 32 if i == 0 else 11
set_row_height(ws, 5, 28)

# CAIXA & RECEBIMENTOS
ws.merge_cells('B6:O6')
write_cell(ws, 6, 2, '  CAIXA & RECEBIMENTOS', orange_subheader())
set_row_height(ws, 6, 24)
write_cell(ws, 7, 2, 'Saldo Inicial de Caixa (R$)', label_cell())
for m in range(12): write_cell(ws, 7, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 7, 15, '=SUM(C7:N7)', total_cell(CURRENCY))
set_row_height(ws, 7, 22)
for i, lbl in enumerate(['Recebimentos B2B (R$)',
                          'Recebimentos Produto/B2C (R$)',
                          'Recebimentos de Editais (R$)',
                          'Outras Receitas (R$)']):
    r = 8 + i
    write_cell(ws, r, 2, lbl, label_cell())
    for m in range(12): write_cell(ws, r, 3 + m, None, input_cell(CURRENCY))
    write_cell(ws, r, 15, f'=SUM(C{r}:N{r})', total_cell(CURRENCY))
    set_row_height(ws, r, 22)
# 12: Recebimentos Totais (auto)
write_cell(ws, 12, 2, 'Recebimentos Totais (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 12, 3 + m, f'=SUM({col}8:{col}11)', auto_cell(CURRENCY))
write_cell(ws, 12, 15, '=SUM(C12:N12)', total_cell(CURRENCY))
set_row_height(ws, 12, 22)

# VENDIDO NO MÊS (cross-sheet)
ws.merge_cells('B14:O14')
write_cell(ws, 14, 2, '  VENDIDO NO MÊS (consolidado das áreas)', orange_subheader())
set_row_height(ws, 14, 24)
# 15: B2B vendido
write_cell(ws, 15, 2, 'Vendido B2B (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 15, 3 + m, f"='🤝 B2B'!{col}18", auto_cell(CURRENCY))
write_cell(ws, 15, 15, '=SUM(C15:N15)', total_cell(CURRENCY))
set_row_height(ws, 15, 22)
# 16: Produto vendido
write_cell(ws, 16, 2, 'Vendido Produto (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 16, 3 + m, f"='🛠 Produto'!{col}18", auto_cell(CURRENCY))
write_cell(ws, 16, 15, '=SUM(C16:N16)', total_cell(CURRENCY))
set_row_height(ws, 16, 22)
# 17: Licitações vendido
write_cell(ws, 17, 2, 'Vendido Licitações (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 17, 3 + m, f"='📋 Licitações'!{col}13", auto_cell(CURRENCY))
write_cell(ws, 17, 15, '=SUM(C17:N17)', total_cell(CURRENCY))
set_row_height(ws, 17, 22)
# 18: Total Vendido
write_cell(ws, 18, 2, 'Total Vendido no Mês (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 18, 3 + m, f'=SUM({col}15:{col}17)', auto_cell(CURRENCY))
write_cell(ws, 18, 15, '=SUM(C18:N18)', total_cell(CURRENCY))
set_row_height(ws, 18, 22)

# DESPESAS
ws.merge_cells('B20:O20')
write_cell(ws, 20, 2, '  DESPESAS', orange_subheader())
set_row_height(ws, 20, 24)
for i, lbl in enumerate(['Folha & RH (R$)',
                          'Marketing/ADS (R$)',
                          'Infra & Tech (R$)',
                          'Terceiros & Freelancers (R$)',
                          'Outras Despesas (R$)']):
    r = 21 + i
    write_cell(ws, r, 2, lbl, label_cell())
    for m in range(12): write_cell(ws, r, 3 + m, None, input_cell(CURRENCY))
    write_cell(ws, r, 15, f'=SUM(C{r}:N{r})', total_cell(CURRENCY))
    set_row_height(ws, r, 22)
# 26: Despesas totais
write_cell(ws, 26, 2, 'Despesas Totais (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 26, 3 + m, f'=SUM({col}21:{col}25)', auto_cell(CURRENCY))
write_cell(ws, 26, 15, '=SUM(C26:N26)', total_cell(CURRENCY))
set_row_height(ws, 26, 22)

# RESULTADO & INDICADORES
ws.merge_cells('B28:O28')
write_cell(ws, 28, 2, '  RESULTADO & INDICADORES', orange_subheader())
set_row_height(ws, 28, 24)
# 29: Resultado = Receb - Despesas
write_cell(ws, 29, 2, 'Resultado do Mês (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 29, 3 + m, f'={col}12-{col}26', auto_cell(CURRENCY))
write_cell(ws, 29, 15, '=SUM(C29:N29)', total_cell(CURRENCY))
set_row_height(ws, 29, 22)
# 30: Burn = IF(Resultado<0, -Resultado, 0)
write_cell(ws, 30, 2, 'Burn Rate (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 30, 3 + m, f'=IF({col}29<0,-{col}29,0)', auto_cell(CURRENCY))
write_cell(ws, 30, 15, '=SUM(C30:N30)', total_cell(CURRENCY))
set_row_height(ws, 30, 22)
# 31: Runway = Saldo Inicial / Burn
write_cell(ws, 31, 2, 'Runway (meses, auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 31, 3 + m, f'=IFERROR({col}7/{col}30,0)', auto_cell(DEC))
write_cell(ws, 31, 15, '=IFERROR(AVERAGE(C31:N31),0)', total_cell(DEC))
set_row_height(ws, 31, 22)
# 32: Margem = Resultado / Recebimentos
write_cell(ws, 32, 2, 'Margem (%, auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 32, 3 + m, f'=IFERROR({col}29/{col}12,0)', auto_cell(PCT))
write_cell(ws, 32, 15, '=IFERROR(AVERAGE(C32:N32),0)', total_cell(PCT))
set_row_height(ws, 32, 22)
# 33: Contas a Receber (input)
write_cell(ws, 33, 2, 'Contas a Receber (R$)', label_cell())
for m in range(12): write_cell(ws, 33, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 33, 15, '=SUM(C33:N33)', total_cell(CURRENCY))
set_row_height(ws, 33, 22)
# 34: Inadimplência
write_cell(ws, 34, 2, 'Inadimplência (R$)', label_cell())
for m in range(12): write_cell(ws, 34, 3 + m, None, input_cell(CURRENCY))
write_cell(ws, 34, 15, '=SUM(C34:N34)', total_cell(CURRENCY))
set_row_height(ws, 34, 22)

# VENDIDO vs RECEBIDO
ws.merge_cells('B36:O36')
write_cell(ws, 36, 2, '  VENDIDO vs RECEBIDO', orange_subheader())
set_row_height(ws, 36, 24)
# 37: Gap (Vendido - Recebido)
write_cell(ws, 37, 2, 'Gap Vendido−Recebido (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 37, 3 + m, f'={col}18-{col}12', auto_cell(CURRENCY))
write_cell(ws, 37, 15, '=SUM(C37:N37)', total_cell(CURRENCY))
set_row_height(ws, 37, 22)
# 38: % conversão
write_cell(ws, 38, 2, '% Conversão Venda→Caixa (auto)', label_cell())
for m in range(12):
    col = get_column_letter(3 + m)
    write_cell(ws, 38, 3 + m, f'=IFERROR({col}12/{col}18,0)', auto_cell(PCT))
write_cell(ws, 38, 15, '=IFERROR(AVERAGE(C38:N38),0)', total_cell(PCT))
set_row_height(ws, 38, 22)

ws.freeze_panes = 'C6'

# =============================================================
# ABA 6 — EXPORT
# =============================================================
ws = wb.create_sheet('📊 Dashboard')
ws.sheet_view.showGridLines = False

# Header
hdrs = ['chave', 'area', 'label'] + MONTHS + ['Total']
for i, h in enumerate(hdrs):
    write_cell(ws, 1, 1 + i, h, {
        'font': Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF'),
        'fill': PatternFill('solid', fgColor=NAVY),
        'alignment': Alignment(horizontal='left' if i < 3 else 'center', vertical='center'),
        'border': box_border,
    })
set_row_height(ws, 1, 28)

# Linhas (mesmas do original)
EXPORT_ROWS = [
    # FINANCEIRO
    ('saldo_inicial',      'Financeiro', 'Saldo Inicial de Caixa', '💰 Financeiro', 7),
    ('receb_b2b',          'Financeiro', 'Recebimentos B2B', '💰 Financeiro', 8),
    ('receb_produto',      'Financeiro', 'Recebimentos Produto/B2C', '💰 Financeiro', 9),
    ('receb_editais',      'Financeiro', 'Recebimentos de Editais', '💰 Financeiro', 10),
    ('receb_outras',       'Financeiro', 'Outras Receitas', '💰 Financeiro', 11),
    ('receb_total',        'Financeiro', 'Recebimentos Totais', '💰 Financeiro', 12),
    ('vendido_b2b',        'Financeiro', 'Vendido B2B', '💰 Financeiro', 15),
    ('vendido_produto',    'Financeiro', 'Vendido Produto', '💰 Financeiro', 16),
    ('vendido_licit',      'Financeiro', 'Vendido Licitações', '💰 Financeiro', 17),
    ('vendido_total',      'Financeiro', 'Total Vendido', '💰 Financeiro', 18),
    ('desp_folha',         'Financeiro', 'Folha & RH', '💰 Financeiro', 21),
    ('desp_ads',           'Financeiro', 'Marketing/ADS', '💰 Financeiro', 22),
    ('desp_tech',          'Financeiro', 'Infra & Tech', '💰 Financeiro', 23),
    ('desp_terceiros',     'Financeiro', 'Terceiros & Freelancers', '💰 Financeiro', 24),
    ('desp_outras',        'Financeiro', 'Outras Despesas', '💰 Financeiro', 25),
    ('desp_total',         'Financeiro', 'Despesas Totais', '💰 Financeiro', 26),
    ('resultado',          'Financeiro', 'Resultado do Mês', '💰 Financeiro', 29),
    ('burn',               'Financeiro', 'Burn Rate', '💰 Financeiro', 30),
    ('runway',             'Financeiro', 'Runway (meses)', '💰 Financeiro', 31),
    ('margem',             'Financeiro', 'Margem (%)', '💰 Financeiro', 32),
    ('contas_receber',     'Financeiro', 'Contas a Receber', '💰 Financeiro', 33),
    ('inadimplencia',      'Financeiro', 'Inadimplência', '💰 Financeiro', 34),
    ('gap_vendido_receb',  'Financeiro', 'Gap Vendido−Recebido', '💰 Financeiro', 37),
    ('conv_caixa',         'Financeiro', '% Conversão Venda→Caixa', '💰 Financeiro', 38),
    # PRODUTO
    ('prod_aulas',         'Produto', 'Aulas/Hot Seats', '🛠 Produto', 7),
    ('prod_marcos',        'Produto', 'Marcos Entregues', '🛠 Produto', 8),
    ('prod_sla',           'Produto', 'SLA de Entrega', '🛠 Produto', 9),
    ('prod_contratos',     'Produto', 'Contratos Ativos', '🛠 Produto', 12),
    ('prod_nps',           'Produto', 'NPS', '🛠 Produto', 13),
    ('prod_49code_qtd',    'Produto', 'Vendas 49 Code (qtd)', '🛠 Produto', 16),
    ('prod_49code_rec',    'Produto', 'Receita 49 Code', '🛠 Produto', 17),
    ('prod_receita',       'Produto', 'Receita Total Produto', '🛠 Produto', 18),
    ('prod_uptime',        'Produto', 'Uptime', '🛠 Produto', 21),
    ('prod_bugs_rep',      'Produto', 'Bugs Reportados', '🛠 Produto', 22),
    ('prod_bugs_res',      'Produto', 'Bugs Resolvidos', '🛠 Produto', 23),
    ('prod_features',      'Produto', 'Features Lançadas', '🛠 Produto', 24),
    # B2B
    ('b2b_ads',            'B2B', 'Investimento ADS', '🤝 B2B', 7),
    ('b2b_leads',          'B2B', 'Leads Gerados', '🤝 B2B', 8),
    ('b2b_mqls',           'B2B', 'MQLs', '🤝 B2B', 9),
    ('b2b_sqls',           'B2B', 'SQLs', '🤝 B2B', 10),
    ('b2b_cac',            'B2B', 'CAC', '🤝 B2B', 11),
    ('b2b_pipeline',       'B2B', 'Pipeline Ativo', '🤝 B2B', 14),
    ('b2b_novos_leads',    'B2B', 'Novos Leads B2B', '🤝 B2B', 15),
    ('b2b_propostas',      'B2B', 'Propostas Enviadas', '🤝 B2B', 16),
    ('b2b_contratos',      'B2B', 'Contratos Fechados', '🤝 B2B', 17),
    ('b2b_receita',        'B2B', 'Receita B2B Vendida', '🤝 B2B', 18),
    ('b2b_ticket',         'B2B', 'Ticket Médio B2B', '🤝 B2B', 19),
    ('b2b_taxa_fech',      'B2B', 'Taxa de Fechamento', '🤝 B2B', 20),
    ('b2b_ciclo',          'B2B', 'Ciclo Médio de Venda', '🤝 B2B', 21),
    # LICITAÇÕES
    ('lic_monit',          'Licitações', 'Editais Monitorados', '📋 Licitações', 7),
    ('lic_subm',           'Licitações', 'Editais Submetidos', '📋 Licitações', 8),
    ('lic_ganhos',         'Licitações', 'Editais Ganhos', '📋 Licitações', 9),
    ('lic_taxa',           'Licitações', 'Taxa de Êxito', '📋 Licitações', 10),
    ('lic_receita',        'Licitações', 'Receita Editais', '📋 Licitações', 13),
    ('lic_pipeline',       'Licitações', 'Pipeline Editais', '📋 Licitações', 14),
    ('lic_vigentes',       'Licitações', 'Contratos Vigentes', '📋 Licitações', 17),
    ('lic_prest_entr',     'Licitações', 'Prestações Entregues', '📋 Licitações', 18),
    ('lic_prest_pend',     'Licitações', 'Prestações Pendentes', '📋 Licitações', 19),
    ('lic_prest_atraso',   'Licitações', 'Prestações em Atraso', '📋 Licitações', 20),
    ('lic_trs',            'Licitações', 'TRs em Aprovação', '📋 Licitações', 23),
    ('lic_followups',      'Licitações', 'Follow-ups Agendados', '📋 Licitações', 24),
]

# Mapeamento de num_format por chave
KEY_FORMATS = {}
for k, _, _, _, _ in EXPORT_ROWS:
    if k.startswith('prod_sla') or k.startswith('prod_uptime'):
        KEY_FORMATS[k] = PCT
    elif k.startswith('lic_taxa') or k.startswith('b2b_taxa') or k == 'margem' or k == 'conv_caixa':
        KEY_FORMATS[k] = PCT
    elif k == 'runway':
        KEY_FORMATS[k] = DEC
    elif k in ('prod_aulas','prod_marcos','prod_contratos','prod_nps','prod_49code_qtd',
               'prod_bugs_rep','prod_bugs_res','prod_features',
               'b2b_leads','b2b_mqls','b2b_sqls','b2b_propostas','b2b_contratos','b2b_novos_leads','b2b_ciclo',
               'lic_monit','lic_subm','lic_ganhos','lic_vigentes','lic_prest_entr','lic_prest_pend','lic_prest_atraso','lic_trs','lic_followups'):
        KEY_FORMATS[k] = INT
    else:
        KEY_FORMATS[k] = CURRENCY

# Render
for i, (key, area, label, sheet, source_row) in enumerate(EXPORT_ROWS):
    r = 2 + i
    # chave / area / label
    write_cell(ws, r, 1, key, {'font': Font(name=FONT_NAME, size=9, bold=True, color=TXT),
                                'alignment': Alignment(horizontal='left', vertical='center', indent=1),
                                'border': box_border})
    write_cell(ws, r, 2, area, {'font': Font(name=FONT_NAME, size=9, color=TXT),
                                 'alignment': Alignment(horizontal='left', vertical='center', indent=1),
                                 'border': box_border})
    write_cell(ws, r, 3, label, {'font': Font(name=FONT_NAME, size=9, color=TXT),
                                  'alignment': Alignment(horizontal='left', vertical='center', indent=1),
                                  'border': box_border})
    fmt = KEY_FORMATS[key]
    # 12 meses + total
    for m in range(12):
        col_in_source = get_column_letter(3 + m)
        formula = f"='{sheet}'!{col_in_source}{source_row}"
        write_cell(ws, r, 4 + m, formula, auto_cell(fmt))
    # Total — col 16
    if fmt == PCT or fmt == DEC:
        write_cell(ws, r, 16, f'=IFERROR(AVERAGE(D{r}:O{r}),0)', total_cell(fmt))
    else:
        write_cell(ws, r, 16, f'=SUM(D{r}:O{r})', total_cell(fmt))
    set_row_height(ws, r, 20)

# Column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 30
for m in range(12):
    ws.column_dimensions[get_column_letter(4 + m)].width = 13
ws.column_dimensions['P'].width = 14

ws.freeze_panes = 'D2'

# =============================================================
# SAVE
# =============================================================
out = '/Users/luanacosta/www/one-page-financeiro-49/Planilha_Resultados_2026.xlsx'
wb.save(out)
print(f'OK — {out}')
print(f'Sheets: {wb.sheetnames}')
